#!/usr/bin/env python3
"""
Generate AEM On-Premise to EDS Migration Estimate (REVISED v3.0)
Samsung Semiconductor + LED Website
Adobe Professional Format

REVISION v3.0 NOTES:
- Incorporates Samsung's internal migration analysis document
- Actual page count: 31,603 pages across 9 regions (not 350-500 estimated)
- Actual component count: 643 components / 132+ types
- 6 custom OSGi bundles, external MySQL, Akamai NetStorage dependencies
- 44 WCM templates + 6 template-types
- Block Library reuse strategy maintained
- Samsung's own assessment: Large category, 9-12 months
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# Adobe brand colors
ADOBE_RED = "FA0F00"
ADOBE_DARK = "2C2C2C"
ADOBE_GRAY = "747474"
ADOBE_LIGHT_GRAY = "F5F5F5"
ADOBE_WHITE = "FFFFFF"
ADOBE_BLUE = "1473E6"
ADOBE_GREEN = "2D9D78"
HEADER_BG = "323232"

# Styles
title_font = Font(name="Adobe Clean", size=18, bold=True, color=ADOBE_DARK)
subtitle_font = Font(name="Adobe Clean", size=14, bold=True, color=ADOBE_GRAY)
header_font = Font(name="Adobe Clean", size=11, bold=True, color=ADOBE_WHITE)
body_font = Font(name="Adobe Clean", size=10, color=ADOBE_DARK)
body_bold_font = Font(name="Adobe Clean", size=10, bold=True, color=ADOBE_DARK)
total_font = Font(name="Adobe Clean", size=11, bold=True, color=ADOBE_RED)
section_font = Font(name="Adobe Clean", size=11, bold=True, color=ADOBE_BLUE)
reuse_font = Font(name="Adobe Clean", size=10, color=ADOBE_GREEN, bold=True)
warning_font = Font(name="Adobe Clean", size=10, bold=True, color="FF6600")

header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
alt_fill = PatternFill(start_color=ADOBE_LIGHT_GRAY, end_color=ADOBE_LIGHT_GRAY, fill_type="solid")
total_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
section_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
reuse_fill = PatternFill(start_color="E8F8F0", end_color="E8F8F0", fill_type="solid")
warning_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_row(ws, row, cols, is_alt=False, is_total=False, is_section=False, is_reuse=False, is_warning=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if is_total:
            cell.font = total_font
            cell.fill = total_fill
        elif is_section:
            cell.font = section_font
            cell.fill = section_fill
        elif is_reuse:
            cell.font = body_font
            cell.fill = reuse_fill
        elif is_warning:
            cell.font = body_font
            cell.fill = warning_fill
        else:
            cell.font = body_font
            if is_alt:
                cell.fill = alt_fill
        if col >= 3:
            cell.alignment = center_align
        else:
            cell.alignment = left_align


def create_workbook():
    wb = Workbook()

    # ============================================================
    # SHEET 1: Executive Summary
    # ============================================================
    ws_exec = wb.active
    ws_exec.title = "Executive Summary"
    ws_exec.sheet_properties.tabColor = ADOBE_RED

    ws_exec.column_dimensions['A'].width = 5
    ws_exec.column_dimensions['B'].width = 85
    ws_exec.column_dimensions['C'].width = 28
    ws_exec.column_dimensions['D'].width = 28

    row = 2
    ws_exec.cell(row=row, column=2, value="ADOBE EXPERIENCE MANAGER").font = Font(name="Adobe Clean", size=10, color=ADOBE_RED, bold=True)
    row += 1
    ws_exec.cell(row=row, column=2, value="Edge Delivery Services Migration Estimate").font = title_font
    row += 1
    ws_exec.cell(row=row, column=2, value="Samsung Semiconductor & LED Division").font = subtitle_font
    row += 1
    ws_exec.cell(row=row, column=2, value="REVISED v3.0 — Based on Samsung Internal Migration Analysis + DAM Report").font = Font(name="Adobe Clean", size=10, italic=True, color=ADOBE_RED)
    row += 2

    # Project Overview
    ws_exec.cell(row=row, column=2, value="PROJECT OVERVIEW").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    overview_items = [
        ("Client:", "Samsung Electronics - Semiconductor Division"),
        ("Source Platform:", "AEM 6.5 On-Premise (Large instance)"),
        ("Target Platform:", "AEM Edge Delivery Services (Cloud)"),
        ("Websites in Scope:", "semiconductor.samsung.com + led.samsung.com"),
        ("Total Pages:", "31,603 pages across 9 regions"),
        ("Total Components:", "643 component nodes / 132+ types"),
        ("DAM Size:", "229.17 GB | 1,692,306 nodes | 8,815,757 properties"),
        ("Languages:", "4 locales (EN, KR, JP, CN) × 9 regions"),
        ("Samsung Classification:", "LARGE (9-12 months standard)"),
        ("Date:", "May 2026"),
        ("Prepared by:", "Adobe Professional Services"),
        ("Version:", "3.0 (Based on Samsung internal analysis document)"),
    ]

    for label, value in overview_items:
        ws_exec.cell(row=row, column=2, value=label).font = body_bold_font
        ws_exec.cell(row=row, column=3, value=value).font = body_font
        row += 1

    row += 2
    ws_exec.cell(row=row, column=2, value="CURRENT STATE — FROM SAMSUNG INTERNAL ANALYSIS").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    # Regional page breakdown
    region_headers = ["Region", "Size", "Pages", "Notes"]
    for col, h in enumerate(region_headers, 2):
        ws_exec.cell(row=row, column=col, value=h)
    for col in range(2, 6):
        cell = ws_exec.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws_exec.column_dimensions['E'].width = 45
    row += 1

    regions = [
        ("global", "347 MB", "7,128", "Global master + multi-region master content"),
        ("kr", "225 MB", "5,674", "Korea (ko_kr)"),
        ("us", "193 MB", "5,351", "US + careers dedicated components"),
        ("emea", "171 MB", "5,041", "Europe + searchjobs dedicated"),
        ("jp", "171 MB", "4,007", "Japan (ja_jp)"),
        ("cn", "155 MB", "4,298", "China (zh_cn)"),
        ("de", "2.2 MB", "33", "Germany (small, insights/newsroom only)"),
        ("ssir", "936 KB", "19", "SSIR dedicated small site"),
        ("ds-test", "6.2 MB", "52", "Test/staging content"),
        ("TOTAL", "1.3 GB", "31,603", ""),
    ]

    for i, (reg, size, pages, notes) in enumerate(regions):
        ws_exec.cell(row=row, column=2, value=reg)
        ws_exec.cell(row=row, column=3, value=size)
        ws_exec.cell(row=row, column=4, value=pages)
        ws_exec.cell(row=row, column=5, value=notes)
        is_t = (i == len(regions) - 1)
        for col in range(2, 6):
            cell = ws_exec.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align if col in [3, 4] else left_align
            if is_t:
                cell.font = total_font
                cell.fill = total_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
                cell.font = body_font
            else:
                cell.font = body_font
        row += 1

    row += 2
    ws_exec.cell(row=row, column=2, value="COMPONENT INVENTORY SUMMARY").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    comp_headers = ["Category", "Types", "Key Components"]
    for col, h in enumerate(comp_headers, 2):
        ws_exec.cell(row=row, column=col, value=h)
    for col in range(2, 5):
        cell = ws_exec.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    components = [
        ("global/content/statics", "44", "hero, accordion, carousel, feature-benefit, table, etc."),
        ("global/content/common", "26", "gnb, footer, breadcrumb, hashtag, contactus, cookie"),
        ("global/content/product", "21", "hero, lnb, spec, product-finder, related-resources"),
        ("global/content/event", "16", "login, mypage, webinar-regist, subscription"),
        ("global/content/article", "11", "article-grid, event-calendar, event-grid, popular-news"),
        ("global/content/search", "4", "search-grid, search-publications"),
        ("global/content/careers", "2", "job-list, job-detail"),
        ("global/page (templates)", "6", "product-page, content-page, empty-page, etc."),
        ("Region-specific", "2", "emea/searchjobs, us/careers"),
        ("TOTAL", "132+ types / 643 nodes", "Including variants"),
    ]

    for i, (cat, types, comps) in enumerate(components):
        ws_exec.cell(row=row, column=2, value=cat)
        ws_exec.cell(row=row, column=3, value=types)
        ws_exec.cell(row=row, column=4, value=comps)
        is_t = (i == len(components) - 1)
        for col in range(2, 5):
            cell = ws_exec.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = left_align
            if is_t:
                cell.font = total_font
                cell.fill = total_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
                cell.font = body_font
            else:
                cell.font = body_font
        row += 1

    row += 2
    ws_exec.cell(row=row, column=2, value="SCOPE SUMMARY (REVISED v3.0)").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    scope_headers = ["Metric", "Semiconductor Only", "Semiconductor + LED"]
    for col, h in enumerate(scope_headers, 2):
        ws_exec.cell(row=row, column=col, value=h)
    for col in range(2, 5):
        cell = ws_exec.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    scope_data = [
        ("Total Pages (confirmed)", "31,603", "31,603 + ~80 LED pages"),
        ("Regions", "9", "9 + 6 LED regional variants"),
        ("Languages/Locales", "4 (EN, KR, JP, CN)", "4"),
        ("Source Components (AEM)", "643 nodes / 132+ types", "643 + LED-specific"),
        ("Target EDS Blocks", "35-40", "45-52"),
        ("  → Reuse from Block Library", "12-15 (style only)", "15-18"),
        ("  → Extend/Customize", "10-12", "12-15"),
        ("  → Custom Development", "10-13", "15-19"),
        ("WCM Templates → EDS Templates", "44 + 6 types → 8-10 EDS", "8-10 + 3 LED"),
        ("OSGi Bundles to Externalize", "6 custom + 3 third-party", "Same (shared)"),
        ("External Integrations", "MySQL, Akamai, Search", "MySQL, Akamai, Search + LED tools"),
        ("Custom Admin UIs", "5 (PIM, privacy, terms, site-ia, tasks)", "Same (shared)"),
        ("DAM Assets", "229 GB / 1.69M nodes", "229 GB / 1.69M nodes"),
        ("Total Effort (Person-Days)", "680-780", "830-950"),
        ("Duration", "9-12 months", "11-14 months"),
    ]

    for i, (metric, semi, both) in enumerate(scope_data):
        ws_exec.cell(row=row, column=2, value=metric)
        ws_exec.cell(row=row, column=3, value=semi)
        ws_exec.cell(row=row, column=4, value=both)
        for col in range(2, 5):
            cell = ws_exec.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_align if col > 2 else left_align
            if i >= len(scope_data) - 2:
                cell.font = total_font
                cell.fill = total_fill
            elif "Reuse" in metric or "Extend" in metric:
                cell.font = body_font
                cell.fill = reuse_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
                cell.font = body_font
            else:
                cell.font = body_font
        row += 1

    row += 2
    ws_exec.cell(row=row, column=2, value="CRITICAL ARCHITECTURE DIFFERENCES: AEM On-Prem vs EDS").font = Font(name="Adobe Clean", size=12, bold=True, color="FF6600")
    row += 2

    arch_notes = [
        "★ AEM 6.5 uses server-side rendering (HTL/Sling) — EDS uses client-side vanilla JS decoration",
        "★ 643 AEM components (HTL) CANNOT be auto-migrated to EDS blocks — must be re-implemented",
        "★ 6 OSGi bundles (Java) have NO equivalent in EDS — must be externalized as API microservices",
        "★ MySQL direct JDBC connection NOT possible in EDS — requires REST API service layer",
        "★ Akamai NetStorage dependency → EDS has its own CDN; assets delivered from edge",
        "★ 5 Custom Admin UIs (PIM, privacy, etc.) → requires separate headless admin application",
        "★ 44 WCM Templates → reduced to 8-10 EDS page templates (document-based authoring)",
        "★ Content migration: 31,603 pages require automated import with per-region validation",
        "",
        "NOTE: AEM Modernization Tools/Agent are designed for AEM 6.x → AEMaaCS (same architecture).",
        "They do NOT apply to EDS migration since EDS is architecturally different (no JCR, no OSGi, no HTL).",
    ]

    for item in arch_notes:
        ws_exec.cell(row=row, column=2, value=item).font = body_font if not item.startswith("★") else warning_font
        row += 1

    # ============================================================
    # SHEET 2: Detailed Estimate - Semiconductor (REVISED v3.0)
    # ============================================================
    ws_semi = wb.create_sheet("Semiconductor - Detailed")
    ws_semi.sheet_properties.tabColor = ADOBE_BLUE

    col_widths = [5, 8, 52, 15, 15, 15, 15, 50]
    for i, w in enumerate(col_widths, 1):
        ws_semi.column_dimensions[get_column_letter(i)].width = w

    row = 2
    ws_semi.cell(row=row, column=2, value="Samsung Semiconductor - EDS Migration (REVISED v3.0)").font = title_font
    row += 1
    ws_semi.cell(row=row, column=2, value="Based on: 31,603 pages | 643 components | 229GB DAM | 9 regions | 6 OSGi bundles").font = subtitle_font
    row += 2

    headers = ["#", "Work Stream", "Task", "Complexity", "Effort (Days)", "Resources", "Notes"]
    for col, h in enumerate(headers, 2):
        ws_semi.cell(row=row, column=col, value=h)
    style_header_row(ws_semi, row, 8)
    row += 1

    semi_tasks = [
        # Discovery
        ("", "DISCOVERY & ARCHITECTURE", "", "", "", "", "", "section"),
        ("1.1", "", "Site Audit — 31,603 pages across 9 regions", "High", "12", "1 Arch + 1 Dev", "Crawl all regions, map content trees", ""),
        ("1.2", "", "DAM Audit — 229GB / 1.69M nodes classification", "High", "10", "1 Arch + 1 Dev", "Active vs archive, duplicate detection", ""),
        ("1.3", "", "Component Mapping (643 → EDS blocks)", "High", "12", "1 Architect", "Map 132+ component types to 35-40 EDS blocks", ""),
        ("1.4", "", "Template Mapping (44 WCM + 6 types → 8-10 EDS)", "High", "6", "1 Architect", "Simplify 50 templates to EDS model", ""),
        ("1.5", "", "Integration Architecture (MySQL, Akamai, APIs)", "High", "10", "1 Arch + 1 Dev", "Externalization strategy for JDBC, NetStorage", "warning"),
        ("1.6", "", "Admin UI Externalization Strategy (5 UIs)", "High", "6", "1 Architect", "PIM, privacy, terms, site-ia, taskmanagement", "warning"),
        ("1.7", "", "Design System Extraction", "Medium", "5", "1 Designer", "Tokens, typography, spacing from 254 CSS files", ""),
        ("1.8", "", "Migration Strategy & Phasing Document", "Medium", "5", "1 PM", "Region-by-region cutover plan", ""),
        ("", "", "Subtotal - Discovery", "", "66", "", "", "total"),

        # Foundation
        ("", "FOUNDATION & SETUP", "", "", "", "", "", "section"),
        ("2.1", "", "EDS Project Scaffolding", "Low", "3", "1 Dev", "Repo setup, CI/CD, environments", ""),
        ("2.2", "", "Global Styles & Design Tokens (from 254 CSS files)", "High", "12", "1 Dev", "Extract from 582 JS / 254 CSS, create design system", ""),
        ("2.3", "", "Header/Navigation Block (cm-semi-gnb)", "High", "12", "1 Dev", "Mega menu, 9 regions, responsive", ""),
        ("2.4", "", "Footer Block (from cm-semi-gnb footer)", "Medium", "5", "1 Dev", "Multi-column, social, legal, regional variants", ""),
        ("2.5", "", "Core Page Templates (8-10 from 44+6 WCM)", "High", "15", "1 Dev", "Product, content, article, event, static, marketing", ""),
        ("2.6", "", "i18n Framework (4 locales × 9 regions)", "High", "12", "1 Dev", "en.json, ko_kr.json, ja_jp.json, zh_cn.json", ""),
        ("2.7", "", "Search Infrastructure (from sr-semi-search-*)", "High", "10", "1 Dev", "Search grid, publications, suggestions", ""),
        ("2.8", "", "Authentication/Login Framework (ev-semi-login)", "High", "8", "1 Dev", "User login, mypage, event registration", "warning"),
        ("", "", "Subtotal - Foundation", "", "77", "", "", "total"),

        # Block Development - REUSE
        ("", "BLOCK DEVELOPMENT — REUSE (AEM Block Library → brand CSS only)", "", "", "", "", "", "section"),
        ("3.1", "", "Hero Block (from st-semi-hero, pd-semi-hero)", "Low", "3", "1 Dev", "REUSE: 2 source components → 1 block + variants", "reuse"),
        ("3.2", "", "Cards Block (multiple variants)", "Low", "3", "1 Dev", "REUSE: Feature cards, product cards", "reuse"),
        ("3.3", "", "Carousel Block (from st-semi carousel)", "Low", "3", "1 Dev", "REUSE: Content carousel, product carousel", "reuse"),
        ("3.4", "", "Accordion Block (from st-semi accordion)", "Low", "2", "1 Dev", "REUSE: FAQ, expandable sections", "reuse"),
        ("3.5", "", "Tabs Block", "Low", "2", "1 Dev", "REUSE: Content switching tabs", "reuse"),
        ("3.6", "", "Video Embed Block", "Low", "2", "1 Dev", "REUSE: YouTube/custom player", "reuse"),
        ("3.7", "", "Table Block (from st-semi table)", "Low", "2", "1 Dev", "REUSE: Data tables, responsive", "reuse"),
        ("3.8", "", "Breadcrumb Block (from cm-semi breadcrumb)", "Low", "1", "1 Dev", "REUSE: Auto-generated from nav", "reuse"),
        ("3.9", "", "CTA Block", "Low", "2", "1 Dev", "REUSE: Call-to-action banners", "reuse"),
        ("3.10", "", "Columns/Grid Block", "Low", "2", "1 Dev", "REUSE: Layout grids", "reuse"),
        ("3.11", "", "Image Gallery Block", "Low", "2", "1 Dev", "REUSE: Lightbox, grid layout", "reuse"),
        ("3.12", "", "Feature-Benefit Block (from st-semi feature-benefit)", "Low", "3", "1 Dev", "REUSE: Product features display", "reuse"),
        ("", "", "Subtotal - Reuse Blocks (12 blocks)", "", "27", "", "Avg 2.3 days/block (brand CSS only)", "total"),

        # Block Development - EXTEND
        ("", "BLOCK DEVELOPMENT — EXTEND (Library + significant customization)", "", "", "", "", "", "section"),
        ("3.13", "", "Contact Form Block (from cm-semi contactus)", "High", "10", "1 Dev", "EXTEND: Multi-step, regional variants, validation", ""),
        ("3.14", "", "Article Grid Block (from ar-semi-article-grid)", "Medium", "6", "1 Dev", "EXTEND: Filters, pagination, sorting", ""),
        ("3.15", "", "Event Calendar Block (from ar-semi event-calendar)", "High", "8", "1 Dev", "EXTEND: Calendar view, event filtering", ""),
        ("3.16", "", "Popular News Block (from ar-semi popular-news)", "Medium", "5", "1 Dev", "EXTEND: Dynamic content, trending", ""),
        ("3.17", "", "Download/Resources Block", "Medium", "5", "1 Dev", "EXTEND: File downloads with filters", ""),
        ("3.18", "", "Cookie/Privacy Block (from cm-semi cookie)", "Medium", "6", "1 Dev", "EXTEND: OneTrust + Samsung privacy policy", ""),
        ("3.19", "", "Hashtag/Tag Filter Block (from cm-semi hashtag)", "Medium", "5", "1 Dev", "EXTEND: Content tagging/filtering", ""),
        ("3.20", "", "Related Resources Block (from pd-semi related-resources)", "Medium", "5", "1 Dev", "EXTEND: Dynamic content recommendations", ""),
        ("3.21", "", "Subscription Block (from ev-semi subscription)", "Medium", "5", "1 Dev", "EXTEND: Newsletter, event signup", ""),
        ("3.22", "", "LNB Block (from pd-semi-lnb)", "Medium", "4", "1 Dev", "EXTEND: Local navigation bar for products", ""),
        ("", "", "Subtotal - Extend Blocks (10 blocks)", "", "59", "", "Avg 5.9 days/block", "total"),

        # Block Development - CUSTOM
        ("", "BLOCK DEVELOPMENT — CUSTOM (No library equivalent, full build)", "", "", "", "", "", "section"),
        ("3.23", "", "Product Spec Block (from pd-semi-spec)", "High", "12", "1 Dev", "CUSTOM: Dynamic spec tables, comparison", ""),
        ("3.24", "", "Product Finder Block (from pd-semi product-finder)", "High", "15", "1 Dev", "CUSTOM: Filterable product catalog, interactive", "warning"),
        ("3.25", "", "Job List/Detail Block (from cr-semi-job-*)", "High", "10", "1 Dev", "CUSTOM: Careers integration, search, apply", ""),
        ("3.26", "", "Webinar Registration Block (from ev-semi webinar-regist)", "High", "8", "1 Dev", "CUSTOM: Registration flow, calendar integration", ""),
        ("3.27", "", "Event Grid Block (from ar-semi event-grid)", "Medium", "6", "1 Dev", "CUSTOM: Events display with filters", ""),
        ("3.28", "", "Search Publications Block (from sr-semi search-publications)", "Medium", "6", "1 Dev", "CUSTOM: Document/whitepaper search", ""),
        ("3.29", "", "MyPage Block (from ev-semi-mypage)", "High", "10", "1 Dev", "CUSTOM: User dashboard, event history", "warning"),
        ("3.30", "", "Searchjobs Block (from emea/searchjobs)", "Medium", "6", "1 Dev", "CUSTOM: Regional job search", ""),
        ("3.31", "", "Regional Contact Tabs Block", "Medium", "5", "1 Dev", "CUSTOM: Region-based content switching", ""),
        ("3.32", "", "Foundry Services Block", "Medium", "5", "1 Dev", "CUSTOM: Process technology showcase", ""),
        ("3.33", "", "Application Showcase Block", "Medium", "5", "1 Dev", "CUSTOM: AI, Server, Auto, Network cards", ""),
        ("", "", "Subtotal - Custom Blocks (11 blocks)", "", "88", "", "Avg 8 days/block", "total"),

        ("", "", "TOTAL BLOCK DEVELOPMENT (33 blocks from 643 components)", "", "174", "", "643 components consolidated into 33 EDS blocks", "total"),

        # External Services (NEW - from OSGi/integrations)
        ("", "EXTERNAL SERVICES & API LAYER (OSGi bundle replacement)", "", "", "", "", "", "section"),
        ("4.1", "", "MySQL REST API Service (replace JDBC connector)", "High", "20", "1 Backend Dev", "★ semi-apiservice externalization, REST endpoints", "warning"),
        ("4.2", "", "Akamai/CDN Migration (replace NetStorageKit)", "High", "10", "1 DevOps", "★ Cloud Manager CDN or direct EDS delivery", "warning"),
        ("4.3", "", "semi-common bundle functions → Edge/Serverless", "High", "15", "1 Dev", "★ 1.7MB core bundle → serverless functions", "warning"),
        ("4.4", "", "semi-product API endpoints", "Medium", "8", "1 Dev", "Product data services", ""),
        ("4.5", "", "semi-article API endpoints", "Medium", "6", "1 Dev", "Article/content services", ""),
        ("4.6", "", "semi-gnb API endpoints (menu data)", "Medium", "5", "1 Dev", "Navigation data service", ""),
        ("4.7", "", "PIM Admin UI (external SPA/headless)", "High", "15", "1 Dev", "★ Product info management app", "warning"),
        ("4.8", "", "Privacy/Terms Admin UI (external)", "Medium", "8", "1 Dev", "Consent & policy management", ""),
        ("4.9", "", "Site-IA & Task Management Admin (external)", "Medium", "8", "1 Dev", "IA management + task workflow", ""),
        ("", "", "Subtotal - External Services", "", "95", "", "Replaces 6 OSGi bundles + 5 admin UIs", "total"),

        # Content Migration - SIGNIFICANTLY INCREASED
        ("", "CONTENT & ASSET MIGRATION (31,603 pages + 229GB DAM)", "", "", "", "", "", "section"),
        ("5.1", "", "Import Script Development (multi-region)", "High", "15", "1 Dev", "Region-aware importer for 9 regions", ""),
        ("5.2", "", "DAM Audit & Cleanup (229GB)", "High", "12", "1 Dev + 1 Content", "Classify 1.69M nodes, remove duplicates/archive", ""),
        ("5.3", "", "DAM Migration Phase 1 — Critical assets (~60GB)", "High", "15", "1 Dev", "Product images, heroes, logos, active media", ""),
        ("5.4", "", "DAM Migration Phase 2 — Supporting assets (~80GB)", "Medium", "12", "1 Dev", "Blog images, event media, documents", ""),
        ("5.5", "", "Asset Optimization & Format Conversion", "Medium", "8", "1 Dev", "WebP/AVIF conversion, compression, CDN setup", ""),
        ("5.6", "", "Content Migration — global region (7,128 pages)", "High", "20", "1 Dev + 2 Content", "Master content + validation", ""),
        ("5.7", "", "Content Migration — kr region (5,674 pages)", "High", "15", "1 Dev + 1 Content", "Korean locale content", ""),
        ("5.8", "", "Content Migration — us region (5,351 pages)", "High", "15", "1 Dev + 1 Content", "US content + careers", ""),
        ("5.9", "", "Content Migration — emea region (5,041 pages)", "High", "14", "1 Dev + 1 Content", "European content + searchjobs", ""),
        ("5.10", "", "Content Migration — jp region (4,007 pages)", "Medium", "12", "1 Dev + 1 Content", "Japanese locale", ""),
        ("5.11", "", "Content Migration — cn region (4,298 pages)", "Medium", "12", "1 Dev + 1 Content", "Chinese locale", ""),
        ("5.12", "", "Content Migration — de/ssir/ds-test (104 pages)", "Low", "3", "1 Content", "Small regions", ""),
        ("5.13", "", "URL Redirect Mapping (31,603 URLs)", "High", "12", "1 Dev", "Bulk 301 redirects, SEO preservation", ""),
        ("5.14", "", "Metadata & SEO Migration", "Medium", "8", "1 Dev", "Schema, OG tags, sitemap for all regions", ""),
        ("", "", "Subtotal - Content & Asset Migration", "", "173", "", "31,603 pages + 229GB assets", "total"),

        # Integrations
        ("", "INTEGRATIONS", "", "", "", "", "", "section"),
        ("6.1", "", "Search Integration (replace sr-semi-search-*)", "High", "12", "1 Dev", "Algolia/Coveo indexing for 31K pages", ""),
        ("6.2", "", "Analytics Setup (Adobe Analytics)", "Medium", "8", "1 Dev", "Event tracking, data layer, 9 regions", ""),
        ("6.3", "", "Form Submission Backend", "Medium", "5", "1 Dev", "API endpoints, notifications", ""),
        ("6.4", "", "CDN & Edge Configuration", "Medium", "5", "1 DevOps", "Caching, headers, security, replace Akamai", ""),
        ("6.5", "", "Event/Webinar Platform Integration", "Medium", "6", "1 Dev", "Registration, calendar sync", ""),
        ("", "", "Subtotal - Integrations", "", "36", "", "", "total"),

        # Testing
        ("", "TESTING & QUALITY ASSURANCE (9 regions × 4 locales)", "", "", "", "", "", "section"),
        ("7.1", "", "Performance Testing & Optimization", "High", "10", "1 Dev", "Lighthouse 95+, Core Web Vitals all templates", ""),
        ("7.2", "", "Accessibility Testing (WCAG 2.1 AA)", "High", "10", "1 QA", "All 33 blocks × accessibility", ""),
        ("7.3", "", "Cross-browser/Device Testing", "Medium", "8", "1 QA", "Chrome, Safari, Firefox, Edge, mobile", ""),
        ("7.4", "", "Content QA — Region-by-region (9 regions)", "High", "25", "2 QA", "Visual regression, links, media per region", ""),
        ("7.5", "", "Asset Integrity Verification", "Medium", "6", "1 QA", "Verify migrated DAM assets render correctly", ""),
        ("7.6", "", "SEO Validation (31K URLs)", "High", "8", "1 Dev", "Rankings preservation, crawl test, redirects", ""),
        ("7.7", "", "Security Testing", "Medium", "5", "1 DevOps", "Headers, CSP, vulnerability scan", ""),
        ("7.8", "", "Integration Testing (MySQL API, Search, Events)", "High", "8", "1 Dev", "End-to-end integration validation", ""),
        ("7.9", "", "UAT Support (region-by-region cutover)", "High", "15", "1 Dev + 1 QA", "Staged: ssir/de → kr/jp → cn → us/emea → global", ""),
        ("", "", "Subtotal - Testing & QA", "", "95", "", "9 regions × 4 locales matrix", "total"),

        # Launch
        ("", "LAUNCH & HANDOVER", "", "", "", "", "", "section"),
        ("8.1", "", "Go-Live Planning — Staged Regional Cutover", "High", "8", "1 PM + 1 Dev", "ssir/de → kr/jp → cn → us/emea → global", ""),
        ("8.2", "", "Author Training & Documentation", "Medium", "8", "1 PM", "EDS authoring guides for 4 locale teams", ""),
        ("8.3", "", "Developer Handover", "Medium", "5", "1 Dev", "Code docs, architecture, API docs", ""),
        ("8.4", "", "Post-Launch Hypercare (4 weeks - staged)", "High", "20", "1 Dev + 1 QA", "Region-by-region monitoring, hotfixes", ""),
        ("", "", "Subtotal - Launch & Handover", "", "41", "", "", "total"),

        # Grand Total
        ("", "GRAND TOTAL — SEMICONDUCTOR ONLY (REVISED v3.0)", "", "", "757", "", "~9-11 months with 5-7 FTEs", "total"),
    ]

    for i, task in enumerate(semi_tasks):
        for col, val in enumerate(task[:7], 2):
            ws_semi.cell(row=row, column=col, value=val)

        tag = task[7]
        style_data_row(ws_semi, row, 8,
                      is_alt=(i % 2 == 0),
                      is_total=(tag == "total"),
                      is_section=(tag == "section"),
                      is_reuse=(tag == "reuse"),
                      is_warning=(tag == "warning"))
        row += 1

    # ============================================================
    # SHEET 3: LED Website (REVISED v3.0)
    # ============================================================
    ws_led = wb.create_sheet("LED Website - Detailed")
    ws_led.sheet_properties.tabColor = "00A86B"

    for i, w in enumerate(col_widths, 1):
        ws_led.column_dimensions[get_column_letter(i)].width = w

    row = 2
    ws_led.cell(row=row, column=2, value="Samsung LED Website - Incremental Effort (REVISED v3.0)").font = title_font
    row += 1
    ws_led.cell(row=row, column=2, value="Incremental when combined with Semiconductor | Shares OSGi/Admin/DAM infrastructure").font = subtitle_font
    row += 2

    headers_led = ["#", "Work Stream", "Task", "Complexity", "Effort (Days)", "Resources", "Notes"]
    for col, h in enumerate(headers_led, 2):
        ws_led.cell(row=row, column=col, value=h)
    style_header_row(ws_led, row, 8)
    row += 1

    led_tasks = [
        ("", "DISCOVERY (INCREMENTAL)", "", "", "", "", "", "section"),
        ("L1.1", "", "LED Site Audit & Inventory (~80 pages)", "Medium", "5", "1 Architect", "Product catalog, 6 regions", ""),
        ("L1.2", "", "LED Template & Block Mapping", "Medium", "3", "1 Architect", "3 additional templates", ""),
        ("L1.3", "", "LED Navigation & IA Design", "Medium", "3", "1 Architect", "Separate nav structure", ""),
        ("", "", "Subtotal - Discovery", "", "11", "", "", "total"),

        ("", "LED BLOCKS — REUSE (Style Only)", "", "", "", "", "", "section"),
        ("L2.1", "", "LED Product Category Block (reuse Cards)", "Low", "3", "1 Dev", "REUSE: Mid/High/CSP/COB/Module cards", "reuse"),
        ("L2.2", "", "LED Application Gallery (reuse Gallery)", "Low", "2", "1 Dev", "REUSE: Horticulture, automotive", "reuse"),
        ("L2.3", "", "Quick Downloads Block (reuse Resources)", "Low", "2", "1 Dev", "REUSE: Datasheet repository", "reuse"),
        ("", "", "Subtotal - LED Reuse Blocks", "", "7", "", "", "total"),

        ("", "LED BLOCKS — EXTEND", "", "", "", "", "", "section"),
        ("L2.4", "", "LED Spec Comparison Block", "Medium", "6", "1 Dev", "EXTEND: Multi-product comparison table", ""),
        ("L2.5", "", "Automotive LED Showcase", "Medium", "4", "1 Dev", "EXTEND: Application-specific display", ""),
        ("L2.6", "", "In-branding Program Block", "Low", "3", "1 Dev", "EXTEND: Partner program info", ""),
        ("L2.7", "", "Sales Network Block", "Medium", "4", "1 Dev", "EXTEND: Regional sales contacts", ""),
        ("", "", "Subtotal - LED Extend Blocks", "", "17", "", "", "total"),

        ("", "LED BLOCKS — CUSTOM", "", "", "", "", "", "section"),
        ("L2.8", "", "LED Component Calculator", "High", "12", "1 Dev", "CUSTOM: Interactive design calculator", "warning"),
        ("L2.9", "", "LED Engine Calculator", "High", "10", "1 Dev", "CUSTOM: Performance modeling tool", "warning"),
        ("L2.10", "", "LED Module Configurator", "High", "8", "1 Dev", "CUSTOM: Product selection wizard", ""),
        ("L2.11", "", "Virtual Exhibition Block", "High", "10", "1 Dev", "CUSTOM: Interactive virtual tour/3D", "warning"),
        ("", "", "Subtotal - LED Custom Blocks", "", "40", "", "", "total"),

        ("", "", "TOTAL LED BLOCK DEVELOPMENT (11 blocks)", "", "64", "", "", "total"),

        ("", "LED CONTENT MIGRATION", "", "", "", "", "", "section"),
        ("L3.1", "", "LED Import Script Customization", "Medium", "5", "1 Dev", "LED-specific parsers", ""),
        ("L3.2", "", "LED Product Pages (20+ pages × 6 regions)", "Medium", "10", "1 Dev + 1 Content", "All product lines, regional variants", ""),
        ("L3.3", "", "LED Application Pages (10+ pages)", "Medium", "5", "1 Content", "Lighting, Auto, Display", ""),
        ("L3.4", "", "LED Support & Tools Pages", "Medium", "4", "1 Content", "Calculators, downloads", ""),
        ("L3.5", "", "LED News & Events", "Low", "3", "1 Content", "Articles, exhibitions", ""),
        ("L3.6", "", "LED Asset Migration", "Medium", "5", "1 Dev", "Product images, datasheets", ""),
        ("L3.7", "", "LED URL Redirects", "Low", "3", "1 Dev", "led.samsung.com mapping", ""),
        ("", "", "Subtotal - Content Migration", "", "35", "", "", "total"),

        ("", "LED INTEGRATIONS & TESTING", "", "", "", "", "", "section"),
        ("L4.1", "", "LED Regional Support (6 regions)", "High", "8", "1 Dev", "America, EMEA, CN, SEA, JP, KR", ""),
        ("L4.2", "", "LED Calculator Validation & Testing", "High", "8", "1 QA", "Calculator accuracy, cross-browser", ""),
        ("L4.3", "", "LED Performance Optimization", "Medium", "4", "1 Dev", "Calculator load performance", ""),
        ("L4.4", "", "LED Content QA (6 regions)", "Medium", "6", "1 QA", "Visual regression per region", ""),
        ("", "", "Subtotal - Integrations & Testing", "", "26", "", "", "total"),

        ("", "GRAND TOTAL — LED INCREMENTAL (REVISED v3.0)", "", "", "136", "", "~5-6 additional weeks", "total"),
    ]

    for i, task in enumerate(led_tasks):
        for col, val in enumerate(task[:7], 2):
            ws_led.cell(row=row, column=col, value=val)

        tag = task[7]
        style_data_row(ws_led, row, 8,
                      is_alt=(i % 2 == 0),
                      is_total=(tag == "total"),
                      is_section=(tag == "section"),
                      is_reuse=(tag == "reuse"),
                      is_warning=(tag == "warning"))
        row += 1

    # ============================================================
    # SHEET 4: Combined Summary
    # ============================================================
    ws_combined = wb.create_sheet("Combined Summary")
    ws_combined.sheet_properties.tabColor = ADOBE_RED

    ws_combined.column_dimensions['A'].width = 5
    ws_combined.column_dimensions['B'].width = 50
    ws_combined.column_dimensions['C'].width = 20
    ws_combined.column_dimensions['D'].width = 20
    ws_combined.column_dimensions['E'].width = 20
    ws_combined.column_dimensions['F'].width = 20

    row = 2
    ws_combined.cell(row=row, column=2, value="Migration Effort Summary — REVISED v3.0 (Samsung Internal Data)").font = title_font
    row += 1
    ws_combined.cell(row=row, column=2, value="31,603 pages | 643 components | 229GB DAM | 9 regions | 6 OSGi bundles").font = subtitle_font
    row += 3

    sum_headers = ["Work Stream", "Semi Only (Days)", "LED Incr. (Days)", "Combined (Days)", "Combined (Weeks)"]
    for col, h in enumerate(sum_headers, 2):
        ws_combined.cell(row=row, column=col, value=h)
    style_header_row(ws_combined, row, 6)
    row += 1

    summary_data = [
        ("Discovery & Architecture", "66", "11", "77", "3.5"),
        ("Foundation & Setup", "77", "—", "77", "3.5"),
        ("Block Development — Reuse (style only)", "27", "7", "34", "1.5"),
        ("Block Development — Extend (customize)", "59", "17", "76", "3.5"),
        ("Block Development — Custom (new build)", "88", "40", "128", "6"),
        ("External Services & API Layer (OSGi replacement)", "95", "—", "95", "4.5"),
        ("Content & Asset Migration (31K pages + 229GB)", "173", "35", "208", "9.5"),
        ("Integrations", "36", "—", "36", "1.5"),
        ("Testing & QA (9 regions × 4 locales)", "95", "26", "121", "5.5"),
        ("Launch & Handover (staged regional cutover)", "41", "—", "41", "2"),
    ]

    for i, (stream, semi, led, combined, weeks) in enumerate(summary_data):
        ws_combined.cell(row=row, column=2, value=stream)
        ws_combined.cell(row=row, column=3, value=semi)
        ws_combined.cell(row=row, column=4, value=led)
        ws_combined.cell(row=row, column=5, value=combined)
        ws_combined.cell(row=row, column=6, value=weeks)
        is_reuse = "Reuse" in stream
        style_data_row(ws_combined, row, 6, is_alt=(i % 2 == 0), is_reuse=is_reuse)
        row += 1

    totals = [
        ("TOTAL — Semiconductor Only", "757", "—", "757", "~9-11 months"),
        ("TOTAL — LED Incremental", "—", "136", "136", "~5-6 weeks"),
        ("TOTAL — Combined (Semi + LED)", "757", "136", "893", "~11-14 months"),
    ]

    for label, semi, led, combined, weeks in totals:
        ws_combined.cell(row=row, column=2, value=label)
        ws_combined.cell(row=row, column=3, value=semi)
        ws_combined.cell(row=row, column=4, value=led)
        ws_combined.cell(row=row, column=5, value=combined)
        ws_combined.cell(row=row, column=6, value=weeks)
        style_data_row(ws_combined, row, 6, is_total=True)
        row += 1

    # Version comparison
    row += 2
    ws_combined.cell(row=row, column=2, value="ESTIMATE EVOLUTION — Why this differs from initial assessment").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    evo_headers = ["Metric", "Initial (Web crawl)", "Samsung Actual", "Impact"]
    for col, h in enumerate(evo_headers, 2):
        ws_combined.cell(row=row, column=col, value=h)
    style_header_row(ws_combined, row, 5)
    row += 1

    evolution = [
        ("Total Pages", "350-500 (estimated)", "31,603 (confirmed)", "90x more pages → content migration 3x effort"),
        ("Components", "25-35 blocks", "643 nodes / 132+ types", "Must consolidate 643→33 EDS blocks"),
        ("Regions", "4 assumed", "9 confirmed", "2.25x validation/testing matrix"),
        ("Templates", "7-9 estimated", "44 WCM + 6 types", "Complex template consolidation effort"),
        ("OSGi Bundles", "Not assessed", "6 custom + 3 third-party", "NEW: API externalization layer needed"),
        ("Admin UIs", "Not assessed", "5 custom UIs", "NEW: Separate headless admin apps"),
        ("External DB", "Not assessed", "MySQL direct JDBC", "NEW: REST microservice required"),
        ("CDN/Storage", "Not assessed", "Akamai NetStorageKit", "NEW: CDN migration/replacement"),
    ]

    for i, (metric, initial, actual, impact) in enumerate(evolution):
        ws_combined.cell(row=row, column=2, value=metric)
        ws_combined.cell(row=row, column=3, value=initial)
        ws_combined.cell(row=row, column=4, value=actual)
        ws_combined.cell(row=row, column=5, value=impact)
        style_data_row(ws_combined, row, 5, is_alt=(i % 2 == 0), is_warning=(i >= 4))
        row += 1

    # Cost
    row += 3
    ws_combined.cell(row=row, column=2, value="COST ESTIMATION (Based on Adobe Professional Services Rates) — REVISED v3.0").font = Font(name="Adobe Clean", size=12, bold=True, color=ADOBE_RED)
    row += 2

    cost_headers = ["Scenario", "Person-Days", "Team Size", "Duration", "Est. Cost Range (USD)"]
    for col, h in enumerate(cost_headers, 2):
        ws_combined.cell(row=row, column=col, value=h)
    style_header_row(ws_combined, row, 6)
    row += 1

    cost_data = [
        ("Semiconductor Only", "757", "5-7 FTEs", "9-11 months", "$1.32M - $1.67M"),
        ("LED Only (standalone*)", "420", "3-4 FTEs", "6-8 months", "$735K - $925K"),
        ("Combined (Semi + LED)", "893", "6-8 FTEs", "11-14 months", "$1.56M - $1.96M"),
    ]

    for i, (scenario, days, team, duration, cost) in enumerate(cost_data):
        ws_combined.cell(row=row, column=2, value=scenario)
        ws_combined.cell(row=row, column=3, value=days)
        ws_combined.cell(row=row, column=4, value=team)
        ws_combined.cell(row=row, column=5, value=duration)
        ws_combined.cell(row=row, column=6, value=cost)
        style_data_row(ws_combined, row, 6, is_alt=(i % 2 == 0))
        row += 1

    row += 2
    ws_combined.cell(row=row, column=2, value="* LED standalone includes shared foundation/API effort absorbed in combined scenario").font = Font(name="Adobe Clean", size=9, italic=True, color=ADOBE_GRAY)
    row += 1
    ws_combined.cell(row=row, column=2, value="Cost based on blended rate of $1,750-$2,200/day for Adobe Professional Services").font = Font(name="Adobe Clean", size=9, italic=True, color=ADOBE_GRAY)
    row += 1
    ws_combined.cell(row=row, column=2, value="Samsung's own assessment: Large category, 9-12 months (aligns with our Semiconductor-only estimate)").font = Font(name="Adobe Clean", size=9, italic=True, color=ADOBE_BLUE)

    # ============================================================
    # SHEET 5: Risk Register (REVISED with Samsung data)
    # ============================================================
    ws_risk = wb.create_sheet("Risk Register")
    ws_risk.sheet_properties.tabColor = "FF6600"

    ws_risk.column_dimensions['A'].width = 5
    ws_risk.column_dimensions['B'].width = 8
    ws_risk.column_dimensions['C'].width = 45
    ws_risk.column_dimensions['D'].width = 12
    ws_risk.column_dimensions['E'].width = 12
    ws_risk.column_dimensions['F'].width = 55
    ws_risk.column_dimensions['G'].width = 20

    row = 2
    ws_risk.cell(row=row, column=2, value="Risk Register (Revised v3.0 — from Samsung analysis)").font = title_font
    row += 3

    risk_headers = ["ID", "Risk Description", "Likelihood", "Impact", "Mitigation Strategy", "Owner"]
    for col, h in enumerate(risk_headers, 2):
        ws_risk.cell(row=row, column=col, value=h)
    style_header_row(ws_risk, row, 7)
    row += 1

    risks = [
        ("R1", "MySQL direct JDBC → Cloud/EDS has no JDBC support", "Confirmed", "High", "Externalize to REST microservice; AEM uses API calls only", "Backend Dev"),
        ("R2", "DAM 224GB single tree (98%) — migration throughput", "High", "High", "CTT chunk split + top-up rounds; pre-cleanup unused assets", "DevOps"),
        ("R3", "Akamai NetStorageKit dependency", "Confirmed", "High", "Migrate to Cloud Manager CDN or EDS native delivery", "DevOps"),
        ("R4", "5 Custom Admin UIs (PIM, privacy, terms, site-ia, tasks)", "Confirmed", "High", "Granite/Coral UI audit; externalize as SPA or headless apps", "Architect"),
        ("R5", "9 regions × 4 locales simultaneous operation", "Confirmed", "High", "Region-by-region cutover; automated validation (Playwright)", "PM"),
        ("R6", "31,603 pages content migration volume", "Confirmed", "High", "Automated import scripts per region; parallel processing", "Content Eng"),
        ("R7", "semi-common 1.7MB core bundle → externalization", "High", "High", "Priority #1 refactoring; AEM Cloud SDK API compatibility matrix", "Tech Lead"),
        ("R8", "643 components consolidated to 33 EDS blocks", "Medium", "High", "POC key blocks in Phase 1; validate mapping assumptions early", "Architect"),
        ("R9", "44 WCM templates + 6 types → 8-10 EDS templates", "Medium", "Medium", "Template consolidation workshop; validate with content teams", "Architect"),
        ("R10", "JSP 9 remaining (admin/taskmanagement)", "Confirmed", "Low", "HTL conversion (small scope) or externalize to admin SPA", "Dev"),
        ("R11", "DAM 1.69M nodes / 8.8M props — index performance", "Medium", "Medium", "Oak Index redesign; metadata cleanup before migration", "DevOps"),
        ("R12", "LED calculator tools complexity", "Medium", "High", "Dedicated spike/POC; third-party widget as fallback", "Tech Lead"),
        ("R13", "SEO ranking impact — 31,603 URL redirects", "High", "High", "Comprehensive redirect map; staged rollout; 90-day monitoring", "SEO"),
        ("R14", "Repository structure not modernized (ui.apps/ui.content)", "Confirmed", "Medium", "Apply Repository Modernizer before migration", "Dev"),
    ]

    for i, (rid, desc, like, impact, mitigation, owner) in enumerate(risks):
        ws_risk.cell(row=row, column=2, value=rid)
        ws_risk.cell(row=row, column=3, value=desc)
        ws_risk.cell(row=row, column=4, value=like)
        ws_risk.cell(row=row, column=5, value=impact)
        ws_risk.cell(row=row, column=6, value=mitigation)
        ws_risk.cell(row=row, column=7, value=owner)
        is_warn = like == "Confirmed"
        style_data_row(ws_risk, row, 7, is_alt=(i % 2 == 0), is_warning=is_warn)
        row += 1

    # ============================================================
    # SHEET 6: Resource Plan
    # ============================================================
    ws_resource = wb.create_sheet("Resource Plan")
    ws_resource.sheet_properties.tabColor = ADOBE_GRAY

    ws_resource.column_dimensions['A'].width = 5
    ws_resource.column_dimensions['B'].width = 25
    ws_resource.column_dimensions['C'].width = 12
    ws_resource.column_dimensions['D'].width = 55
    ws_resource.column_dimensions['E'].width = 22

    row = 2
    ws_resource.cell(row=row, column=2, value="Recommended Team Structure (Combined Semi + LED)").font = title_font
    row += 3

    res_headers = ["Role", "Count", "Responsibilities", "Duration"]
    for col, h in enumerate(res_headers, 2):
        ws_resource.cell(row=row, column=col, value=h)
    style_header_row(ws_resource, row, 5)
    row += 1

    resources = [
        ("Solution Architect", "1", "Component mapping (643→33), template strategy, integration architecture", "Full duration"),
        ("Senior EDS Developer", "2", "Block development (extend/custom), EDS patterns, performance", "Full duration"),
        ("Backend Developer", "1", "MySQL API service, OSGi externalization, serverless functions", "Months 2-8"),
        ("Frontend Developer", "1-2", "Block styling (reuse tier), 254 CSS migration, responsive, a11y", "Months 2-10"),
        ("Content/DAM Engineer", "1-2", "DAM audit (229GB), 31K page migration, import scripts, per-region QA", "Months 4-12"),
        ("QA Engineer", "1-2", "9-region testing, accessibility, cross-browser, integration tests", "Months 6-14"),
        ("Project Manager", "1", "Region cutover planning, stakeholder coordination, training", "Full duration"),
        ("UX Designer", "0.5", "Design system extraction from 582 JS / 254 CSS", "Months 1-3"),
        ("DevOps Engineer", "1", "CDN migration, Akamai replacement, CI/CD, DAM migration tooling", "Months 2-6, 10-14"),
    ]

    for i, (role, count, resp, duration) in enumerate(resources):
        ws_resource.cell(row=row, column=2, value=role)
        ws_resource.cell(row=row, column=3, value=count)
        ws_resource.cell(row=row, column=4, value=resp)
        ws_resource.cell(row=row, column=5, value=duration)
        style_data_row(ws_resource, row, 5, is_alt=(i % 2 == 0))
        row += 1

    row += 3
    ws_resource.cell(row=row, column=2, value="Total Team: 6-8 FTEs at peak (months 4-10)").font = total_font
    row += 1
    ws_resource.cell(row=row, column=2, value="Ramp: Start with 4, scale to 8 at month 4, taper to 4 from month 10").font = body_font

    # ============================================================
    # SHEET 7: Timeline
    # ============================================================
    ws_timeline = wb.create_sheet("Timeline")
    ws_timeline.sheet_properties.tabColor = "9B59B6"

    ws_timeline.column_dimensions['A'].width = 5
    ws_timeline.column_dimensions['B'].width = 40
    for i in range(3, 60):
        ws_timeline.column_dimensions[get_column_letter(i)].width = 3

    row = 2
    ws_timeline.cell(row=row, column=2, value="Project Timeline — Combined (REVISED v3.0)").font = title_font
    row += 1
    ws_timeline.cell(row=row, column=2, value="11-14 months | Staged regional cutover").font = subtitle_font
    row += 2

    # Month headers (14 months)
    ws_timeline.cell(row=row, column=2, value="Phase / Activity")
    months = ["M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12", "M13", "M14"]
    for m_idx, m in enumerate(months):
        ws_timeline.cell(row=row, column=m_idx + 3, value=m)
    style_header_row(ws_timeline, row, 16)
    row += 1

    phases = [
        ("Phase 1: Discovery & Architecture", 1, 2),
        ("Phase 2: Foundation & Setup", 2, 4),
        ("Phase 3a: Block Reuse (style only)", 3, 4),
        ("Phase 3b: Block Extend (customize)", 3, 6),
        ("Phase 3c: Block Custom (Semi + LED)", 4, 9),
        ("Phase 4: External Services / API Layer", 3, 8),
        ("Phase 5a: DAM Audit & Asset Migration", 4, 8),
        ("Phase 5b: Content Migration (31K pages)", 5, 11),
        ("Phase 6: Integrations", 6, 8),
        ("Phase 7: Testing & QA (per region)", 7, 12),
        ("Phase 8: UAT — ssir/de cutover", 9, 10),
        ("Phase 9: UAT — kr/jp cutover", 10, 11),
        ("Phase 10: UAT — cn/us/emea cutover", 11, 12),
        ("Phase 11: UAT — global cutover + hypercare", 12, 14),
    ]

    colors = ["4A90D9", "2ECC71", ADOBE_GREEN, "27AE60", "E74C3C", "FF6600",
              "9B59B6", "F39C12", "1ABC9C", "3498DB", "E91E63", "C0392B", "8E44AD", ADOBE_RED]

    for i, (phase, start, end) in enumerate(phases):
        ws_timeline.cell(row=row, column=2, value=phase).font = body_bold_font
        ws_timeline.cell(row=row, column=2).border = thin_border
        for m in range(1, 15):
            cell = ws_timeline.cell(row=row, column=m + 2)
            cell.border = thin_border
            if start <= m <= end:
                cell.fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
        row += 1

    row += 2
    ws_timeline.cell(row=row, column=2, value="Staged Regional Cutover Strategy (from Samsung analysis):").font = body_bold_font
    row += 1
    ws_timeline.cell(row=row, column=2, value="ssir/de (small, low risk) → kr/jp → cn → us/emea → global (master, highest risk last)").font = body_font

    # Save
    output_path = "/backups/riteskum/lge-be-eds/repo/Samsung_EDS_Migration_Estimate.xlsx"
    wb.save(output_path)
    print(f"Revised v3.0 estimate saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    create_workbook()
